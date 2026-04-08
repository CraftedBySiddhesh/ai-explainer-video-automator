#!/usr/bin/env python3
"""
AI Explainer Video Prompt Generator - XLSX Multi-Sheet Output
Each section gets its own sheet with Character Sheets table and Scenes table
"""

import sys
import json
import os
import re
from pathlib import Path
from dotenv import load_dotenv
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Load environment variables
load_dotenv()


def call_claude_api(prompt, system_prompt=""):
    """Make API call to Claude"""
    try:
        api_key = os.getenv('ANTHROPIC_API_KEY')
        if not api_key:
            print("❌ Error: ANTHROPIC_API_KEY not found in .env file")
            return None

        url = "https://api.anthropic.com/v1/messages"
        headers = {
            "Content-Type": "application/json",
            "x-api-key": api_key,
            "anthropic-version": "2023-06-01"
        }

        messages = [{"role": "user", "content": prompt}]

        payload = {
            "model": "claude-sonnet-4-20250514",
            "max_tokens": 8000,
            "messages": messages
        }

        if system_prompt:
            payload["system"] = system_prompt

        response = requests.post(url, headers=headers, json=payload)
        response.raise_for_status()

        data = response.json()
        return data["content"][0]["text"]

    except Exception as e:
        print(f"❌ API Error: {e}")
        return None


def clean_json_response(response):
    """Clean model response before JSON parsing"""
    if not response:
        return None

    response = response.strip()
    if response.startswith('```json'):
        response = response[7:]
    if response.startswith('```'):
        response = response[3:]
    if response.endswith('```'):
        response = response[:-3]
    return response.strip()


def sanitize_character_name(name):
    """Force character names to a single clean word"""
    if not name:
        return "Character"

    # Keep letters/numbers/spaces/hyphens only
    name = re.sub(r'[^A-Za-z0-9\s-]', '', str(name)).strip()

    # Split into words and keep first meaningful one
    parts = [p for p in re.split(r'[\s_-]+', name) if p]
    if not parts:
        return "Character"

    cleaned = parts[0].capitalize()

    # Avoid generic empty-ish values
    if cleaned.lower() in {"the", "a", "an", "and"}:
        return "Character"

    return cleaned


def sanitize_text_field(text):
    """Normalize whitespace and remove unnecessary extra text"""
    if not text:
        return ""
    text = str(text).replace("\n", " ")
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def remove_scene_prefix(text):
    """Remove Scene 1:, Scene 2:, etc from prompts"""
    if not text:
        return ""
    text = re.sub(r'^\s*Scene\s*\d+\s*:\s*', '', str(text), flags=re.IGNORECASE).strip()
    return text


def remove_text_rendering_phrases(text):
    """Remove phrases that may cause text to appear inside generated images"""
    if not text:
        return ""

    banned_patterns = [
        r'\bwith text\b',
        r'\binclude text\b',
        r'\bcontaining text\b',
        r'\bshowing text\b',
        r'\blabel(?:s|led)?\b',
        r'\bcaption(?:s|ed)?\b',
        r'\btitle(?:s|d)?\b',
        r'\blogo(?:s)?\b',
        r'\bsign(?:s)?\b',
        r'\bposter(?:s)?\b',
        r'\bbanner(?:s)?\b',
        r'\bletter(?:s)?\b',
        r'\bword(?:s)?\b',
        r'\btypography\b',
        r'\bwatermark(?:s)?\b',
    ]

    cleaned = str(text)
    for pattern in banned_patterns:
        cleaned = re.sub(pattern, '', cleaned, flags=re.IGNORECASE)

    cleaned = re.sub(r'\s+,', ',', cleaned)
    cleaned = re.sub(r'\s+\.', '.', cleaned)
    cleaned = re.sub(r'\s{2,}', ' ', cleaned).strip(" ,.-")
    return cleaned


def enforce_image_prompt_rules(text):
    """Enforce concise flat 2D minimal illustration rules for image prompts"""
    text = sanitize_text_field(text)
    text = remove_scene_prefix(text)
    text = remove_text_rendering_phrases(text)

    required_style = (
        "Flat 2D minimalistic illustration, clean vector shapes, white background, "
        "no text, no labels, no typography, simple composition."
    )

    # Remove duplicated style phrases if model already included them loosely
    lower_text = text.lower()
    if "flat 2d" in lower_text or "minimalistic" in lower_text or "no text" in lower_text:
        final = text
    else:
        final = f"{required_style} {text}"

    final = sanitize_text_field(final)
    return final


def enforce_animation_prompt_rules(text):
    """Keep animation prompts concise and remove scene prefixes"""
    text = sanitize_text_field(text)
    text = remove_scene_prefix(text)
    return text


def sanitize_character_description(text):
    """Ensure character image descriptions are text-free and stylistically consistent"""
    text = sanitize_text_field(text)
    text = remove_text_rendering_phrases(text)

    style_requirements = (
        "Flat 2D minimalistic illustration. Black stick figure with only essential colored clothing or accessories. "
        "White background. No text, no labels, no typography. Show 3-4 directions."
    )

    lower_text = text.lower()
    if "flat 2d" in lower_text or "no text" in lower_text:
        final = text
    else:
        final = f"{style_requirements} {text}"

    final = sanitize_text_field(final)
    return final


def post_process_prompt_data(prompt_data):
    """Normalize model output while preserving output structure"""
    if not prompt_data:
        return prompt_data

    # Character sheets
    character_sheets = prompt_data.get("character_sheets", [])
    cleaned_characters = []
    name_map = {}

    for char in character_sheets:
        original_name = sanitize_text_field(char.get("name", "Character"))
        cleaned_name = sanitize_character_name(original_name)
        cleaned_desc = sanitize_character_description(char.get("description", ""))

        name_map[original_name] = cleaned_name
        cleaned_characters.append({
            "name": cleaned_name,
            "description": cleaned_desc
        })

    prompt_data["character_sheets"] = cleaned_characters

    # Scenes
    cleaned_scenes = []
    for scene in prompt_data.get("scenes", []):
        image_prompt = enforce_image_prompt_rules(scene.get("image_prompt", ""))
        animation_prompt = enforce_animation_prompt_rules(scene.get("animation_prompt", ""))

        # Replace old multi-word character names in image prompts with sanitized single-word names
        for old_name, new_name in sorted(name_map.items(), key=lambda x: len(x[0]), reverse=True):
            if old_name and new_name:
                image_prompt = re.sub(rf'\b{re.escape(old_name)}\b', new_name, image_prompt)

        cleaned_scene = {
            "scene_number": scene.get("scene_number", ""),
            "scene_start": sanitize_text_field(scene.get("scene_start", "")),
            "description": sanitize_text_field(scene.get("description", "")),
            "image_prompt": sanitize_text_field(image_prompt),
            "animation_prompt": sanitize_text_field(animation_prompt),
        }
        cleaned_scenes.append(cleaned_scene)

    prompt_data["scenes"] = cleaned_scenes
    return prompt_data


def detect_sections_with_embeddings(script_text):
    """Use AI to detect sections based on semantic understanding"""

    system_prompt = """You are an expert content analyst. Analyze educational scripts and detect distinct narrative sections based on semantic meaning and topic shifts.
You MUST respond ONLY with valid JSON. No markdown. No commentary."""

    prompt = f"""Analyze this complete educational script and break it into distinct sections.

Read the ENTIRE script first. Identify where topics change, where new stories/civilizations begin.
Each section should be a complete narrative arc about ONE topic/civilization/event.

FULL SCRIPT:
{script_text}

Use semantic understanding to detect section boundaries - NOT just formatting.
Detect topic shifts, story changes, new civilizations, different time periods.

OUTPUT FORMAT (JSON ONLY - NO OTHER TEXT):
{{
  "total_sections": <number>,
  "sections": [
    {{
      "section_number": 1,
      "title": "Extracted or inferred title for this section",
      "content": "Complete text content of this section",
      "summary": "One sentence describing what this section covers"
    }}
  ]
}}

RESPOND WITH JSON ONLY. NO MARKDOWN FENCES."""

    print("🧠 Using AI to analyze script and detect sections...")
    response = call_claude_api(prompt, system_prompt)

    if not response:
        return None

    response = clean_json_response(response)

    try:
        data = json.loads(response)
        return data
    except json.JSONDecodeError as e:
        print(f"  ❌ JSON Parse Error: {e}")
        return None


def generate_section_prompts(section_num, section_data, full_script_context):
    """Generate prompts with Scene_Start mapped to exact script text"""

    system_prompt = """You are an expert AI explainer video prompt generator.
You create flat 2D minimalistic illustrative educational visuals.

STRICT RULES:
- Respond ONLY with valid JSON
- Character names must be ONE WORD ONLY
- Character images must contain NO TEXT
- Image prompts must request NO TEXT, NO LABELS, NO TYPOGRAPHY
- Keep wording concise and avoid unnecessary extra text
- Visual style must be flat 2D, minimalistic, illustrative
- Do not add scene prefixes like "Scene 1:"
"""

    prompt = f"""Generate complete explainer video prompts for this section.

FULL SCRIPT CONTEXT:
{full_script_context[:2000]}...

CURRENT SECTION {section_num}: {section_data['title']}
SECTION SUMMARY: {section_data['summary']}

SECTION CONTENT:
{section_data['content']}

CRITICAL RULES:
- Determine optimal scene count between 14-25 scenes
- For EACH scene, extract the EXACT starting text from the script (Scene_Start)
- Scene_Start should be the exact sentence/phrase where that scene begins in the narration
- Identify all unique character types (1-15 depending on content)
- Character names must be SINGLE WORD ONLY
- Create character sheets as flat 2D minimalistic stick figures
- Character image descriptions must contain NO TEXT, NO LABELS, NO TYPOGRAPHY
- Image prompts must contain NO TEXT, NO LABELS, NO TYPOGRAPHY
- Add only essential wording; no unnecessary extra text

IMAGE PROMPTS MUST:
- Be 35-50 words
- Reference character names from character sheets
- Use ONLY single-word character names
- NOT start with "Scene 1:", "Scene 2:" etc
- Describe composition, character positions, 2-color palette, key visual elements
- Specify flat 2D minimalistic illustrative style
- Explicitly avoid text in the image
- Keep phrasing concise

ANIMATION PROMPTS MUST:
- Be 35-50 words
- NOT start with "Scene 1:", "Scene 2:" etc
- Describe camera movement type, character animations, pacing
- Include specific motion details: camera direction, character gestures, timing
- Keep phrasing concise

OUTPUT FORMAT (JSON ONLY):
{{
  "section_number": {section_num},
  "section_title": "{section_data['title']}",
  "total_scenes": <number between 14-25>,
  "character_sheets": [
    {{
      "name": "SingleWordName",
      "description": "Flat 2D minimalistic illustration. Simple black stick figure with only essential colored clothing or accessories. White background. No text. Show in 3-4 directions."
    }}
  ],
  "scenes": [
    {{
      "scene_number": 1,
      "scene_start": "EXACT text from script where this scene begins",
      "description": "Brief scene description",
      "image_prompt": "35-50 words. Flat 2D minimalistic illustration, 2-color palette, clean composition, single-word character name, no text, no labels, no typography.",
      "animation_prompt": "35-50 words. Camera movement, character motion, pacing, concise motion details."
    }}
  ]
}}

REMEMBER:
- Character names from character sheets ONLY
- Character names must be one word only
- No scene prefixes
- No unnecessary extra text
- No text inside generated images

RESPOND WITH JSON ONLY."""

    print(f"  → Generating detailed prompts for Section {section_num}...")
    response = call_claude_api(prompt, system_prompt)

    if not response:
        return None

    response = clean_json_response(response)

    try:
        data = json.loads(response)
        data = post_process_prompt_data(data)
        return data
    except json.JSONDecodeError as e:
        print(f"  ❌ JSON Parse Error: {e}")
        return None


def create_xlsx_with_sheets(output_path, all_sections_data):
    """Create XLSX with one sheet per section, 2 tables per sheet"""

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Header styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for section_data in all_sections_data:
        section_num = section_data['section_number']

        # Create sheet for this section
        ws = wb.create_sheet(title=f"Section {section_num}")

        # TABLE 1: CHARACTER SHEETS
        ws['A1'] = "CHARACTER SHEETS"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:B1')

        # Character sheets headers
        ws['A2'] = "Character_Name"
        ws['B2'] = "Description"

        ws['A2'].fill = header_fill
        ws['A2'].font = header_font
        ws['A2'].alignment = header_alignment
        ws['B2'].fill = header_fill
        ws['B2'].font = header_font
        ws['B2'].alignment = header_alignment

        # Character sheets data
        char_row = 3
        for char in section_data.get('character_sheets', []):
            ws[f'A{char_row}'] = char['name']
            ws[f'B{char_row}'] = char['description']
            ws[f'B{char_row}'].alignment = Alignment(wrap_text=True)
            char_row += 1

        # Empty rows between tables
        scene_start_row = char_row + 2

        # TABLE 2: SCENES
        ws[f'A{scene_start_row}'] = "SCENES"
        ws[f'A{scene_start_row}'].font = Font(bold=True, size=14)
        ws.merge_cells(f'A{scene_start_row}:E{scene_start_row}')

        # Scene headers
        header_row = scene_start_row + 1
        ws[f'A{header_row}'] = "Scene_No"
        ws[f'B{header_row}'] = "Scene_Start"
        ws[f'C{header_row}'] = "Scene_Description"
        ws[f'D{header_row}'] = "Image_Prompt"
        ws[f'E{header_row}'] = "Animation_Prompt"

        for col in ['A', 'B', 'C', 'D', 'E']:
            cell = ws[f'{col}{header_row}']
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Scene data
        data_row = header_row + 1
        for scene in section_data.get('scenes', []):
            ws[f'A{data_row}'] = scene['scene_number']
            ws[f'B{data_row}'] = scene.get('scene_start', '')
            ws[f'C{data_row}'] = scene.get('description', '')
            ws[f'D{data_row}'] = scene['image_prompt']
            ws[f'E{data_row}'] = scene['animation_prompt']

            # Wrap text for readability
            for col in ['B', 'C', 'D', 'E']:
                ws[f'{col}{data_row}'].alignment = Alignment(wrap_text=True, vertical='top')

            data_row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 50

    wb.save(output_path)
    print(f"✓ XLSX created: {output_path}")


def main():
    if len(sys.argv) < 2:
        print("Usage: python generate_video_prompts.py <script.txt>")
        sys.exit(1)

    script_file = Path(sys.argv[1])

    if not script_file.exists():
        print(f"❌ File not found: {script_file}")
        sys.exit(1)

    print(f"📖 Reading script: {script_file}")
    script_text = script_file.read_text(encoding='utf-8')

    # Detect sections
    section_data = detect_sections_with_embeddings(script_text)

    if not section_data:
        print("❌ Failed to detect sections")
        sys.exit(1)

    sections = section_data['sections']
    total_sections = section_data['total_sections']

    print(f"✓ Detected {total_sections} sections\n")

    # Generate prompts for each section
    all_sections_data = []

    for section in sections:
        section_num = section['section_number']
        print(f"🎬 Processing Section {section_num}: {section['title'][:50]}...")

        prompt_data = generate_section_prompts(section_num, section, script_text)

        if not prompt_data:
            print(f"  ⚠️  Skipping section {section_num}")
            continue

        print(f"  ✓ Generated {prompt_data.get('total_scenes', 0)} scenes")
        all_sections_data.append(prompt_data)
        print()

    # Create XLSX with multiple sheets
    output_file = script_file.stem + '_prompts.xlsx'
    output_path = Path(output_file)

    print(f"💾 Creating XLSX with {len(all_sections_data)} sheets...")
    create_xlsx_with_sheets(output_path, all_sections_data)

    print(f"\n✅ SUCCESS!")
    print(f"✓ Total sections: {len(all_sections_data)}")
    print(f"✓ Total sheets: {len(all_sections_data)}")
    print(f"✓ Output file: {output_path}")


if __name__ == '__main__':
    main()
